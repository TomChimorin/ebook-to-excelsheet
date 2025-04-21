from bs4 import BeautifulSoup
import openpyxl
import pyinputplus as pyip
import urllib.request
import datetime
import re

# Function to extract book names, authors, narrators, and publish year from HTML content
def readLibrary(websiteHTML):
    
    nowTime = datetime.datetime.now()
    # Create BeautifulSoup object to parse HTML
    beautifulSoup = BeautifulSoup(websiteHTML, "html.parser")

    # Regular expression pattern to match years in creator elements
    yearPattern = re.compile(r"(\d{4})")

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    rowIndex = 2  # Start from row 2 to skip the header

    # Write the header
    sheet.cell(row=1, column=1, value="Book Names")
    sheet.cell(row=1, column=2, value="Author")
    sheet.cell(row=1, column=3, value="Narrator")
    sheet.cell(row=1, column=4, value="Publish Year")

    # Find all book elements
    books = beautifulSoup.find('div', class_='resultsContainer')

    # Iterate through each book element
    for book in books.find_all('div',class_='title-result-row'):
       
        bookName = book.find('h3',class_='title-result-row__title').text.strip()
        print(bookName)
        # Get the book name
        

        # Initialize variables to store current author and narrator 
        currentAuthor = ""
        currentNarrator = ""

        # Find the corresponding creator elements for the book  
        creatorElement= book.find('h3',class_='title-result-row__creator').text
    
        year = "2024"
        # Iterate through each creator 
        creatorElement = creatorElement.split("Author")
        currentAuthor = creatorElement[0]
        
           # Initialize year variable outside the loop
        for creator in creatorElement:
                if "Narrator" in creator:
                    # Search for the year pattern in the creator string
                    year_match = re.search(yearPattern, creator)  
                    if year_match:
                        # Get the matched year with brackets
                        year_with_brackets = year_match.group(0)
                        # Extract only the year without brackets
                        year = year_match.group(1)  
                     # Remove the year with brackets from the narrator string
                        currentNarrator = creatorElement[creatorElement.index(creator)].replace("Narrator", "").replace(year_with_brackets, "").replace("()","").strip()  
                    else:
                        # Remove only the "Narrator" tag
                        currentNarrator = creatorElement[creatorElement.index(creator)].replace("Narrator", "").strip()  
                      
                    
                else:
                    currentNarrator = "None"

                   
        
        # Write book details to the Excel sheet
        sheet.cell(row=rowIndex, column=1, value=bookName)
        sheet.cell(row=rowIndex, column=2, value=currentAuthor)
        sheet.cell(row=rowIndex, column=3, value=currentNarrator)
        sheet.cell(row=rowIndex, column=4, value=year)

        # Increment row index for next iteration
        rowIndex += 1

    # Save the Excel file
    workbook.save("library.xlsx")
    print("Printed Successfully")
    print("Start Time = ", nowTime)
    endTime = datetime.datetime.now()
    print("Finish printing time =", endTime)

# List of URLs to choose from
URLList = [
    "https://www.overdrive.com/collections/1122780/biggest-books-of-march",
    "https://www.overdrive.com/collections/1135883/biggest-books-of-april",
    "https://www.overdrive.com/collections/1131528/biggest-books-of-spring-2024",
    "https://www.overdrive.com/collections/45038/top-2023-releases",
]

# List of options corresponding to URLs
urlChoices = [
    'Biggest Books of March',
    'Biggest Books of April',
    'Biggest Books of Spring 2024',
    "Top 2023 Releases",
]

# Retry loop to handle exceptions during URL retrieval
retryCount = 0
while retryCount < 3:
    try:
        # Prompt user to select a URL
        urlChoice = pyip.inputMenu(list(urlChoices), prompt="Which link would you like to visit in the library?\n", numbered=True)

        # Open the selected URL
        libraryWebsite = urllib.request.urlopen(URLList[urlChoices.index(urlChoice)])
        websiteHTML = libraryWebsite.read()
        print("year")
        # Process HTML content to extract book details
        readLibrary(websiteHTML)
        
        break  # Break the loop if successful
    except pyip.TimeoutException:
        print("Timeout: No input provided.")
    except pyip.RetryLimitException:
        print("Retry limit exceeded. Exiting.")
        break  # Break the loop if retry limit exceeded
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print("Error accessing the website:", e)
    except Exception as e:
        print("An error occurred:", e)
    retryCount += 1
 
else:
    print("Failed to retrieve website after 3 attempts.")
       
